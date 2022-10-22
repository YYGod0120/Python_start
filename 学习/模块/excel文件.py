from openpyxl import Workbook,load_workbook
wb = load_workbook('专业群号.xlsx')

print(wb.sheetnames)

sheet = wb["Sheet1"]

# a0 = sheet['B2:B14']
# print(a0)
# for cell in a0:
#     print(int(cell[0].value)) 获取某一切片内的元素

#按行遍历全页并且更改
for row in sheet:
    for cell in row:
        #print(cell.coordinate) #单元格号
        if cell.value == None:
            n = cell.coordinate
            sheet[n] = 123456
        else:
            print(cell.value,end=',')
    print()

#指定第几行到第几行
for row in sheet.iter_rows(min_row=2,max_row=5,max_col=5):
    for cell in row:
        print(cell.value,end=",")
    print()



#按列遍历全页
for colnum in sheet.columns:
    for cell in colnum:
        print(cell.value,end=',')
    print()

#指定第几列到第几列
for col in sheet.iter_cols(min_col=2,max_col=5,):
    for i in col:
        print(i.value,end=",")
    print()



#删除
# wb.remove(sheet)
# del wb[sheet]

#保存
wb.save('专业群号.xlsx')